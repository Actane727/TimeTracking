#!/usr/bin/python3

from openpyxl import*
import datetime as dt
import warnings, os
from time import sleep
from threading import Thread

currentyr = dt.datetime.now().strftime('%y')
filepath = "/home/pi/Desktop/TrainingTime '" + currentyr + ".xlsx"
grouppath = "/home/pi/Desktop/GroupTraining '" + currentyr + ".xlsx"
logs = 0
oldtime = 0
persold = 0

if os.path.isfile(filepath):
    pass
else:
    wb = Workbook()
    wb.save(filepath)

if os.path.isfile(grouppath):
    pass
else:
    book = Workbook()
    book.save(grouppath)

warnings.simplefilter("ignore")
wb = load_workbook(filepath)
warnings.simplefilter("default")
sheets = wb.sheetnames

if 'Sheet' in sheets:
            pass
else:
            wb.create_sheet('Sheet', 0)

ws = wb['Sheet']


def clearsave():
    global wb, ws, condition
    if condition == 0:
        ws.title = dt.date.today().strftime('%b %d, %Y')
        wb.create_sheet('Sheet', 0)
        ws = wb['Sheet']
        wb.save(filepath)
        condition = 1
    else:
        print("/" * 79)
        print("**The saved document is already at the most current document. Saving now will create a blank document.")
        print("/" * 79 + "\n")


def justsave():
    print("The file has been saved at " + dt.datetime.now().strftime('%m/%d/%y %H:%M') + ".\n")
    wb.save(filepath)


def oneshot():
    global currentyr, wb, filepath
    run_once = 0
    while True:
        if run_once == 0:
            if dt.datetime.now().strftime('%a %H %P') == "Sun 01 am":
                print("/" * 79)
                print("         **The new sheet has been made and is ready for the new week!**")
                print("/" * 79 + "\n")
                clearsave()
                if dt.datetime.now().strftime('%y') == currentyr:
                    pass
                else:
                    os.system('sudo mkdir /home/pi/Desktop/TimeFiles' + currentyr)
                    os.system('sudo mv /home/pi/Desktop/*.xlsx /home/pi/Desktop/TimeFiles' + currentyr)
                    currentyr = dt.datetime.now().strftime('%y')
                    wb = Workbook()
                    wb.save(filepath)
                    print("          **A new workbook has been created for the new year!**\n")
                print("Please swipe your card to sign in to the Maintenance Department Training Room.")
                print("-" * 14 + "Type \"save\" to quick save or \"reboot\" to reboot Pi" + "-" * 14 + "\n")
                run_once = 1
            else:
                run_once = 0


def restart():
    wb.save(filepath)
    answer = input("Are you sure that you want to reboot now? y/n ")
    if answer == "n":
        print("The system will not shutdown.\n")
        pass
    elif answer == "y":
        print("-" * 79 + "\nThe system will reboot now...\n")
        sleep(2)
        print("Rebooting.........")
        sleep(3)
        os.system('sudo shutdown -r now')
    else:
        print("That is not a proper answer...\n")
        pass


def save():
    global condition
    print("-" * 79 + "\nThe file has been saved at " + dt.datetime.now().strftime('%D %H:%M') + ".\n")
    wb.save(filepath)
    condition = 0


def train():
    global topic, grouppath
    training = True
    oneline = True
    oldnumber = 0
    book = load_workbook(grouppath)
    names = book.sheetnames

    if topic in names:
        sheet = book[topic]
        pass
    else:
        if 'Sheet' in names:
            sheet = book['Sheet']
            sheet.title = topic
            book.create_sheet('Sheet', 0)
            book.save(grouppath)
            sheet = book[topic]
        else:
            book.create_sheet('Sheet', 0)
            sheet = book['Sheet']
            sheet.title = topic
            book.create_sheet('Sheet', 0)
            book.save(grouppath)
            sheet = book[topic]

    while training:
        if oneline:
            print('Please swipe your card to sign in to {} Training.'.format(topic))
            oneline = False
        number = input()
        if number == "0900704147":
            book.save(grouppath)
            print("Group Training Stopped!!!\n")
            training = False
            break
        else:
            try:
                int(number)
                pass
            except ValueError:
                print(number + " is not a valid personnel number!\n")
                continue
        if number == oldnumber:
            deltat = dt.datetime.now() - firsttime
            if deltat.total_seconds() <= 300:
                continue
            else:
                pass
        else:
            pass
        next_row = sheet.max_row + 1
        sheet['B{}'.format(next_row)] = int(number)
        sheet['C{}'.format(next_row)] = dt.datetime.now().strftime('%-m/%-d/%Y %l:%M:00 %p')
        firsttime = dt.datetime.now()
        oldnumber = number


def main():
    global condition, logs, oldtime, persold, persnum, topic
    while True:
        print("Please swipe your card to sign in to the Maintenance Department Training Room.")
        print("-" * 14 + "Type \"save\" to quick save or \"reboot\" to reboot Pi" + "-" * 14 + "\n")
        persnum = input()
        if persnum == "reboot":
            restart()
            continue
        elif persnum == "save":
            justsave()
            continue
        elif persnum == "0900878466":
            topic = "Confined Space"
            train()
            continue
        elif persnum == "0900558136":
            topic = "Fire Safety"
            train()
            continue
        elif persnum == "0900762450":
            topic = "Hazardous Energy"
            train()
            continue
        elif persnum == "0900550943":
            topic = "Other"
            train()
            continue
        elif persnum == "0900824101":
            topic = "HazMat"
            train()
            continue
        elif persnum == "0900704147":
            continue
        else:
            try:
                int(persnum)
                pass
            except ValueError:
                print(persnum + " is not a valid personnel number!\n")
                continue
        if persnum == persold:
            diff = dt.datetime.now() - oldtime
            if diff.total_seconds() <= 300:
                print("You have just clocked in " + str(int(diff.total_seconds())) + " seconds ago.\n")
                continue
            else: pass
        else: pass
        input_row = wb.active.max_row + 1
        ws['B{}'.format(input_row)] = int(persnum)
        ws['C{}'.format(input_row)] = dt.datetime.now().strftime('%-m/%-d/%y %l:%M:00 %p')
        print("Hello, your personnel number is " + str(persnum))
        if logs >= 5:
            print("Your data is saved on row " + str(input_row), "at " +
                  str(dt.datetime.now().strftime('%D %l:%M %p')) + ".")
        else:
            print("Your data is saved on row " + str(input_row), "at " +
                  str(dt.datetime.now().strftime('%D %l:%M %p')) + ".\n")
        logs += 1
        condition = 0
        oldtime = dt.datetime.now()
        persold = persnum
        if logs >= 5:
            save()
            logs = 0


if __name__ == "__main__":
    t1 = Thread(target = oneshot)
    t2 = Thread(target = main)
    t1.setDaemon(True)
    t2.setDaemon(True)
    t1.start()
    t2.start()
    t1.join()
    while True:
        pass
